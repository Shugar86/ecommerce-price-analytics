"""
Прайс Syperopt (XLSX).
"""

from __future__ import annotations

import logging
from io import BytesIO
from typing import Any, Optional

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.collectors.normalized_io import replace_normalized_offers, upsert_source_health
from app.collectors.xls_common import (
    first_barcode,
    guess_vendor_code,
    normalize_vendor_code,
    parse_price_ru,
)

logger = logging.getLogger(__name__)

SYPEROPT_XLSX_URL = (
    "http://www.syperopt.ru/price_wago_abb_legrand_iek_495t5890043_syperopt_ru.xlsx"
)


def _find_header_row_openpyxl(ws: Worksheet) -> tuple[Optional[int], dict[str, int]]:
    """Первая строка заголовков (первые 50) и {lower: col_1based}."""
    max_r = min(50, ws.max_row or 0)
    max_c = ws.max_column or 0
    for r in range(1, max_r + 1):
        values: list[str] = []
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            values.append(str(v).strip() if v is not None else "")
        joined = " ".join(x.lower() for x in values if x)
        if any(
            k in joined
            for k in ("артик", "наимен", "цена", "штрих", "barcode", "ean", "код", "name")
        ):
            header_map: dict[str, int] = {}
            for c, v in enumerate(values, start=1):
                key = v.strip().lower()
                if key:
                    header_map[key] = c
            return r, header_map
    return None, {}


def _map_cols(header_map: dict[str, int]) -> tuple[Optional[int], ...]:
    """Возвращает 1-based индексы колонок: name, price, vendor, barcode, brand."""

    def find_col(*needles: str) -> Optional[int]:
        for k, idx in header_map.items():
            for n in needles:
                if n in k:
                    return idx
        return None

    return (
        find_col("наимен", "товар", "номенклат", "product", "name", "наим", "модел"),
        find_col("цена", "price"),
        find_col("артик", "код", "sku", "vendor", "арт.", "art"),
        find_col("штрих", "barcode", "ean", "gtin"),
        find_col("бренд", "brand", "произв"),
    )


def iter_syperopt_rows(ws: Worksheet) -> list[dict[str, Any]]:
    """Читает все строки XLSX в формате для normalized_offers (публичный API)."""
    return _iter_syperopt_rows_impl(ws)


def _guess_brand_from_syperopt_name(name: str) -> str | None:
    """Эвристика бренда по шапке наименования (практика Syperopt)."""
    n = name.upper()
    if "WAGO" in n:
        return "WAGO"
    if "LEGRAND" in n:
        return "Legrand"
    if "SCHNEIDER" in n:
        return "Schneider Electric"
    if " ИЭК" in name or "IEK" in n:
        return "IEK"
    if n.startswith("ABB") or " ABB " in f" {n} ":
        return "ABB"
    return None


def _iter_syperopt_rows_impl(ws: Worksheet) -> list[dict[str, Any]]:
    """Реализация: сначала TDM-подобные заголовки, иначе фикс. колонки B=наименование, D=цена."""
    out: list[dict[str, Any]] = []
    header_r, hmap = _find_header_row_openpyxl(ws)
    c_name: int | None = None
    c_price: int | None = None
    c_vendor: int | None = None
    c_barcode: int | None = None
    c_brand: int | None = None
    start_r = 1
    if header_r is not None and hmap:
        m = _map_cols(hmap)
        c_name, c_price, c_vendor, c_barcode, c_brand = m[0], m[1], m[2], m[3], m[4]
        if c_name and c_price:
            start_r = header_r + 1
        else:
            c_name = c_price = None

    use_fixed_bd = c_name is None or c_price is None
    if use_fixed_bd:
        # Прайс syperopt.ru: наименование в B (2), первая цена с НДС в D (4).
        c_name, c_price, c_vendor, c_barcode, c_brand = 2, 4, None, None, None
        start_r = 1
        logger.info("Syperopt: фиксированные колонки B/D, полный iter_rows (read-friendly)")

    if use_fixed_bd:
        # read_only: max_row ненадёжен — итерируем весь лист
        for r, row in enumerate(
            ws.iter_rows(min_row=1, values_only=True), 1
        ):
            if not row or len(row) < 4:
                continue
            name = str((row[1] if len(row) > 1 else "") or "").strip()
            if not name or name.lower() in ("nan", "none", "итого"):
                continue
            if len(name) < 5:
                continue
            pv = row[3] if len(row) > 3 else None
            if pv in (None, ""):
                continue
            try:
                if isinstance(pv, (int, float)):
                    price_rub = float(pv)  # type: ignore[arg-type]
                else:
                    price_rub = parse_price_ru(str(pv))
            except ValueError:
                continue
            if price_rub <= 0:
                continue
            vendor = guess_vendor_code(name)
            barcode = first_barcode(name)
            brand = _guess_brand_from_syperopt_name(name)
            out.append(
                {
                    "name": name,
                    "price_rub": price_rub,
                    "vendor_code": vendor,
                    "barcode": barcode,
                    "brand": brand,
                    "external_id": f"syperopt_{r}",
                }
            )
        return out

    for r in range(start_r, (ws.max_row or 0) + 1):
        def cell(col: int | None) -> str:
            if col is None:
                return ""
            v = ws.cell(r, col).value
            if v is None:
                return ""
            return str(v).strip()

        name = cell(c_name)
        if not name or name.lower() in ("nan", "none", "итого"):
            continue
        if len(name) < 5:
            continue
        pv = ws.cell(r, c_price).value if c_price is not None else None
        if pv in (None, ""):
            continue
        try:
            if isinstance(pv, (int, float)):
                price_rub = float(pv)  # type: ignore[arg-type]
            else:
                price_rub = parse_price_ru(str(pv))
        except ValueError:
            continue
        if price_rub <= 0:
            continue

        vendor = (
            normalize_vendor_code(cell(c_vendor)) if c_vendor else None
        ) or guess_vendor_code(name)
        bc_raw = cell(c_barcode) if c_barcode else ""
        barcode = first_barcode(bc_raw) if bc_raw else first_barcode(name)
        brand = cell(c_brand) if c_brand else _guess_brand_from_syperopt_name(name)

        out.append(
            {
                "name": name,
                "price_rub": price_rub,
                "vendor_code": vendor,
                "barcode": barcode,
                "brand": brand,
                "external_id": f"syperopt_{r}",
            }
        )
    return out


def fetch_syperopt_offers(session: Session) -> None:
    """Скачивает XLSX Syperopt и пишет в normalized_offers."""
    try:
        logger.info("Syperopt: загрузка XLSX")
        response = requests.get(
            SYPEROPT_XLSX_URL,
            timeout=(15, 600),
            headers={"User-Agent": "PriceDesk-Collector/1.0"},
        )
        response.raise_for_status()
        buf = BytesIO(response.content)
        wb = load_workbook(buf, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return
            rows = iter_syperopt_rows(ws)
        finally:
            wb.close()
        source_name = "Syperopt XLSX"
        replace_normalized_offers(
            session, source_name, SYPEROPT_XLSX_URL, rows, loaded_at=None
        )
        upsert_source_health(
            session, source_name, SYPEROPT_XLSX_URL, rows
        )
        session.commit()
        logger.info("Syperopt: сохранено %s строк", len(rows))
    except requests.RequestException as e:
        logger.error("Syperopt: HTTP: %s", e)
        session.rollback()
    except (OSError, ValueError) as e:
        logger.error("Syperopt: файл: %s", e)
        session.rollback()
    except SQLAlchemyError as e:
        logger.error("Syperopt: БД: %s", e)
        session.rollback()
    except Exception as e:
        logger.error("Syperopt: %s", e)
        session.rollback()
